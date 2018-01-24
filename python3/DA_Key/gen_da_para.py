
fs = ['da_boards-1.key', 'da_boards-2.key','da_boards-3.key','da_boards-4.key','da_boards-6.key']
ftarget = 'da_para.txt'
para = []

name = ''
gain = ''
offset = ''

for f in fs:
    fr = open(f, 'r')
    lines = fr.readlines()
    for line in lines:
        if line.find('"name"') > 0:
            name = str.strip(line[line.index(':')+1:line.index('",')+1])
            name = name[1:-1]
        if line.find('"gain"') > 0:
            gain = str.strip(line[line.index(':')+1:line.index(']')+1])
        if line.find('offsetCorr') > 0:
            offset = str.strip(line[line.index(':')+1:line.index(']')+1])
            para.append([name, gain, offset])


import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
import random

file_name = '两次测试对比.xlsx'
wb = load_workbook(file_name)
sheet = wb.get_sheet_by_name('Sheet1')
i = 0
for row in range(1, sheet.max_row+1):
    if sheet.cell(row=row, column=2).value == 'ch1':
        if i > 0:
            para.append([name, gain, offset])
        name = sheet.cell(row=row, column=1).value
    if sheet.cell(row=row, column=1).value == '增益':
        tmp = [str(sheet.cell(row=row, column=2).value), str(sheet.cell(row=row, column=3).value),str(sheet.cell(row=row, column=4).value),str(sheet.cell(row=row, column=5).value),]
        gain = '['+','.join(tmp) + ']'
    if sheet.cell(row=row, column=1).value == '偏置':
        tmp = [str(sheet.cell(row=row, column=2).value), str(sheet.cell(row=row, column=3).value),str(sheet.cell(row=row, column=4).value),str(sheet.cell(row=row, column=5).value),]
        offset = '['+','.join(tmp) + ']'
    i += 1
wb.close()
fw = open(ftarget, 'w')

new_l = para.sort()
for row in sorted(para):
    # print(':'.join(row))
    fw.write(':'.join(row)+'\n')
fw.close()
