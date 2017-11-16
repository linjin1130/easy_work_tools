import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

result_zy_file = "F:\\PycharmProjects\\easy_work_tools\\自由空间商务车-不能报.xlsx"
result_xc_file = "F:\\PycharmProjects\\easy_work_tools\\自由空间小车-不能报.xlsx"
result_sw_file = "F:\\PycharmProjects\\easy_work_tools\\教授用车-不能报.xlsx"

result_zy_file_can = "F:\\PycharmProjects\\easy_work_tools\\自由空间商务车-能报.xlsx"
result_xc_file_can = "F:\\PycharmProjects\\easy_work_tools\\自由空间小车-能报.xlsx"
result_sw_file_can = "F:\\PycharmProjects\\easy_work_tools\\教授用车-能报.xlsx"

freespace_people_file = "F:\\PycharmProjects\\easy_work_tools\\自由空间通讯录.xlsx"
def read_people(file_name):
    wb = load_workbook(file_name)
    print(wb.sheetnames)
    total_people = []

    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    print(file_name, sheet.max_column, sheet.max_row)
    for i in range(2,sheet.max_row+1):
        value = sheet.cell(row=i, column=1).value
        if value:
            total_people.append(value.replace(' ', ''))
    print(len(total_people))
    wb.close()
    return total_people

freespace_people = set(read_people(freespace_people_file))
print(freespace_people)

os.chdir('F:\\PycharmProjects\\easy_work_tools\\python3\\原始数据-分车型和自由空间')
print(os.getcwd())

file_list_can = []
file_list_canot = []
for dirname, dirshere, fileshere in os.walk(os.getcwd()):
    # print("dirname", dirname)
    # print("dirshere", dirshere)
    # print("fileshere", fileshere)
    for name in fileshere:
        if name.endswith('.xlsx'):
            if(dirname.find('不能')<0):
                file_list_can.append(os.path.join(dirname,name))
            else:
                file_list_canot.append(os.path.join(dirname, name))

print(file_list_can[0])

def replace_fun(str_in):
    str_out = str_in
    if str_in == '李扬':
        str_out = '李杨'
    if str_in == '陈夏薇':
        str_out = '陈厦微'
    if str_in == '谢鸿波':
        str_out = '谢虹波'
    if str_in == '王伟阳':
        str_out = '王卫阳'
    if str_in == '李雪娇':
        str_out = '李雪姣'
    if str_in == '曹源':
        str_out = '曹原'
    if str_in == '蔡文琦':
        str_out = '蔡文奇'
    if str_in == '李正平':
        str_out = '黎正平'
    if str_in == '印亚运':
        str_out = '印亚云'
    if str_in == '谢洪波':
        str_out = '谢虹波'
    if str_in == '阎锦智':
        str_out = '闫锦智'
    if str_in == '戴飞':
        str_out = '戴辉'
    if str_in == '叶巍巍':
        str_out = '叶魏魏'
    if str_in == '刘蔚悦':
        str_out = '刘尉悦'
    if str_in == u'徐煜':
        str_out = '徐昱'
    if str_in == '龚应洪':
        str_out = '龚云洪'
    if str_in == '钟先峰':
        str_out = '钟先锋'
    if str_in == '李海滨':
        str_out = '李海兵'
    if str_in == '陈夏微':
        str_out = '陈厦微'
    if str_in == '李光兵':
        str_out = '李广兵'
    if str_in == '曹雷':
        str_out = '曹蕾'
    if str_in == '徐煜':
        str_out = '徐昱'
    if str_in == ' 谢鸿波':
        str_out = '谢虹波'
    if str_in == '彭老师':
        str_out = '彭承志'
    if str_in == '应娟':
        str_out = '印娟'
    return str_out

import datetime
def read_file(file_name):
    wb = load_workbook(file_name)
    print(wb.sheetnames)
    total_data = []

    #. 读出sheet，
    sw = []
    xc = []
    zy = []

    for sheet_name in wb.sheetnames:
        print(sheet_name)
        sheet = wb.get_sheet_by_name(sheet_name)
        print(file_name, sheet.max_column, sheet.max_row)
        # print(sheet.title)

        for i in range(1,sheet.max_row+1):
            row = []
            for j in range(1,sheet.max_column):
                row.append(sheet.cell(row=i, column=j).value)

            # print(type(row[0]), row)
            if row[0] :
                tmp= replace_fun(row[5])
                row[5] = tmp

                row_0 = str(row[0])
                row[0] = row_0
                if row_0.find('日期') > -1:
                    print("首行", row)
                elif row_0.find('车费') < 0:
                    if sheet_name.find('商务')>-1:
                        sw.append(row)
                    elif sheet_name.find('小车')>-1:
                        xc.append(row)
                    elif sheet_name.find('自由')>-1:
                        zy.append(row)
                else:
                    print("ddddd", row)
                # else:
                #     if sheet_name.find('商务')>-1:
                #         sw.append(row)
                #     elif sheet_name.find('小车')>-1:
                #         xc.append(row)
                #     elif sheet_name.find('自由')>-1:
                #         zy.append(row)
    wb.close()
    return sw, xc, zy


shangwu_l = []
xiaoche_l = []
freespace_l = []

sum_= [0,0,0]

for file in file_list_can:
    sw, xc, zy = read_file(file)
    shangwu_l.extend(sw)
    xiaoche_l.extend(xc)
    freespace_l.extend(zy)

name_idx = 5

print("自由空间清单")
for row in freespace_l:
    if row[name_idx] in freespace_people:
        pass
    else:
        print("教授", row[name_idx])
        shangwu_l.append(row)
        freespace_l.remove(row)

    print(row)
    sum_[0] += row[name_idx-1]


print("商务车清单")
for row in shangwu_l:
    if row[name_idx] in freespace_people:
        print("自由空间", row[name_idx])
        freespace_l.append(row)
        shangwu_l.remove(row)

    try:
        sum_[1] += int(row[name_idx-1])
    except:
        print(row)

print("小车清单")
for row in xiaoche_l:
    if row[name_idx] in freespace_people:
        pass
    else:
        print("教授", row[name_idx])
        shangwu_l.append(row)
        xiaoche_l.remove(row)
    sum_[2] += row[name_idx-1]
print(sum_, sum(sum_))

def gen_dic(key_idx, data_l):
    dic = {}
    for row in data_l:
        if key_idx == 0:
            key = row[key_idx][:row[key_idx].rfind('.')]
        else:
            key = row[key_idx]
        if key in dic.keys():
            dic[key].append(row)
        else:
            dic[key] = [row]
    return dic

dic_zy = gen_dic(0, freespace_l)
dic_sw = gen_dic(name_idx, shangwu_l)
dic_xc = gen_dic(0, xiaoche_l)

def gen_results(dic, path, name_idx):
    wb = Workbook()
    total = []
    for key in sorted(dic.keys()):
        sum_ = 0
        print(key)
        print(dic[key])
        ws = wb.create_sheet(title=key)
        for row in range(1, len(dic[key])+1):
            for col in range(1,len(dic[key][row-1])+1):
                ws.cell(column=col, row=row, value=dic[key][row-1][col-1])
            sum_ += dic[key][row-1][name_idx-1]
        ws.cell(column=1, row=len(dic[key])+3, value="车费(元）")
        ws.cell(column=2, row=len(dic[key]) + 3, value=sum_)
        ws.cell(column=3, row=len(dic[key]) + 3, value='税（车费*5%）：')
        ws.cell(column=4, row=len(dic[key]) + 3, value=sum_*0.05)
        ws.cell(column=5, row=len(dic[key]) + 3, value='含税总计')
        ws.cell(column=6, row=len(dic[key]) + 3, value=sum_*1.05)
        total.append([key, sum_, sum_*1.05])
    ws = wb.get_sheet_by_name('Sheet')
    total_sum = 0
    for row in range(len(total)):
        for col in range(len(total[row])):
            ws.cell(column=col+1, row=row+2, value=total[row][col])
        total_sum += total[row][1]


    ws.cell(column=1, row=1, value='项目')
    ws.cell(column=2, row=1, value='总计')
    ws.cell(column=3, row=1, value='含税总计')

    ws.cell(column=1, row=len(total) + 3, value='总计')
    ws.cell(column=2, row=len(total) + 3, value=total_sum)
    ws.cell(column=3, row=len(total) + 3, value=total_sum*1.05)
    wb.save(path)

gen_results(dic_zy, result_zy_file_can, name_idx)
gen_results(dic_sw, result_sw_file_can, name_idx)
gen_results(dic_xc, result_xc_file_can, name_idx)

# gen_results(dic_zy, result_zy_file)
# gen_results(dic_sw, result_sw_file)
# gen_results(dic_xc, result_xc_file)
# for ii in dic_zy.keys():
#     print(ii)
#     for jj in dic_zy[ii]:
#         print(jj)
#
# for ii in dic_xc.keys():
#     print(ii)
#     for jj in dic_xc[ii]:
#         print(jj)
#
# for ii in dic_sw.keys():
#     print(ii)
#     for jj in dic_sw[ii]:
#         print(jj)