import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
import random

from openpyxl.utils import get_column_letter

result_zy_file = "F:\\PycharmProjects\\easy_work_tools\\自由空间用车-不能报.xlsx"
result_xc_file = "F:\\PycharmProjects\\easy_work_tools\\自由空间小车-不能报.xlsx"
result_sw_file = "F:\\PycharmProjects\\easy_work_tools\\教授用车-不能报.xlsx"

result_ce_file = "F:\\PycharmProjects\\easy_work_tools\\核减差额.xlsx"

result_zy_file_can = "F:\\PycharmProjects\\easy_work_tools\\自由空间用车-能报.xlsx"
result_xc_file_can = "F:\\PycharmProjects\\easy_work_tools\\自由空间小车-能报.xlsx"
result_sw_file_can = "F:\\PycharmProjects\\easy_work_tools\\教授用车-能报.xlsx"

result_cya_file_can = "F:\\PycharmProjects\\easy_work_tools\\陈宇翱-用车-能报.xlsx"
result_cya_file = "F:\\PycharmProjects\\easy_work_tools\\陈宇翱-用车-不能报.xlsx"

freespace_people_file = "F:\\PycharmProjects\\easy_work_tools\\自由空间通讯录.xlsx"

total_money = 0

def read_people(file_name):
    wb = load_workbook(file_name)
    print("原始数据sheet名字：")
    print(wb.sheetnames)
    total_people = []

    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    print(file_name, sheet.max_column, sheet.max_row)
    for i in range(2,sheet.max_row+1):
        value = sheet.cell(row=i, column=1).value
        if value:
            total_people.append(value.replace(' ', ''))
    print('总人数',len(total_people))
    wb.close()
    return total_people

def replace_time(in_time):
    new_l = in_time.split('.')
    if int(new_l[-1]) == 1:
        new_l[-1] = '01'
    if int(new_l[-1]) == 2:
        new_l[-1] = '02'
    if int(new_l[-1]) == 3:
        new_l[-1] = '03'
    if int(new_l[-1]) == 4:
        new_l[-1] = '04'
    if int(new_l[-1]) == 5:
        new_l[-1] = '05'
    if int(new_l[-1]) == 6:
        new_l[-1] = '06'
    if int(new_l[-1]) == 7:
        new_l[-1] = '07'
    if int(new_l[-1]) == 8:
        new_l[-1] = '08'
    if int(new_l[-1]) == 9:
        new_l[-1] = '09'

    new_str = '.'.join(new_l)
    return new_str

freespace_people = set(read_people(freespace_people_file))
print('自由空间人员：')
print(freespace_people)

professor_list = set(['霍永恒','张强','逯鹤','陈帅','陆朝阳','陈宇翱'])

chepai = ['沪B：HR569','沪F：57920','沪B：ZB860','苏J：1016E','苏A：M676E' ]
chepai_sw = ['沪B：TJ908','沪F：FZ127']
head_title = ['日期', '时间', '人数', '起点至终点', '金额', '用车人', '备注', '车型', '车牌']
os.chdir('F:\\PycharmProjects\\easy_work_tools\\python3\\张梅3')
print('当前目录：',os.getcwd())

file_list_can = []
file_list_canot = []
for dirname, dirshere, fileshere in os.walk(os.getcwd()):
    # print("dirname", dirname)
    # print("dirshere", dirshere)
    # print("fileshere", fileshere)
    for name in fileshere:
        if name.endswith('.xlsx'):
            if(dirname.find('不能')<0):
                file_list_can.append(os.path.join(dirname,name))
            else:
                file_list_canot.append(os.path.join(dirname, name))

def replace_fun(str_in):
    str_out = str_in
    if str_in == '李扬':
        str_out = '李杨'
    if str_in == '陈夏薇':
        str_out = '陈厦微'
    if str_in == '谢鸿波':
        str_out = '谢虹波'
    if str_in == '王伟阳':
        str_out = '王卫阳'
    if str_in == '李雪娇':
        str_out = '李雪姣'
    if str_in == '曹源':
        str_out = '曹原'
    if str_in == '蔡文琦':
        str_out = '蔡文奇'
    if str_in == '李正平':
        str_out = '黎正平'
    if str_in == '印亚运':
        str_out = '印亚云'
    if str_in == '谢洪波':
        str_out = '谢虹波'
    if str_in == '阎锦智':
        str_out = '闫锦智'
    if str_in == '戴飞':
        str_out = '戴辉'
    if str_in == '叶巍巍':
        str_out = '叶魏魏'
    if str_in == '刘蔚悦':
        str_out = '刘尉悦'
    if str_in == u'徐煜':
        str_out = '徐昱'
    if str_in == '龚应洪':
        str_out = '龚云洪'
    if str_in == '钟先峰':
        str_out = '钟先锋'
    if str_in == '李海滨':
        str_out = '李海兵'
    if str_in == '陈夏微':
        str_out = '陈厦微'
    if str_in == '李光兵':
        str_out = '李广兵'
    if str_in == '曹雷':
        str_out = '曹蕾'
    if str_in == '徐煜':
        str_out = '徐昱'
    if str_in == ' 谢鸿波':
        str_out = '谢虹波'
    if str_in == '彭老师':
        str_out = '彭承志'
    if str_in == '应娟':
        str_out = '印娟'
    if str_in == '朱小波':
        str_out = '朱晓波'
    if str_in == '李利':
        str_out = '李力'
    if str_in == '丁少华老师':
        str_out = '丁少华'
    if str_in == '朱燕南':
        str_out = '朱燕楠'
    if str_in == '路朝阳':
        str_out = '陆朝阳'
    return str_out

import datetime
def read_file(file_name):
    wb = load_workbook(file_name)
    print('文件名：',wb.sheetnames)
    total_data = []

    #. 读出sheet，
    sw = []
    xc = []
    zy = []

    for sheet_name in wb.sheetnames:
        print('文件中的sheet名',sheet_name)
        sheet = wb.get_sheet_by_name(sheet_name)
        print(file_name, sheet.max_column, sheet.max_row)
        # print(sheet.title)
        active_num = 0
        for i in range(1,sheet.max_row+1):
            row = []
            for j in range(1,10):
                row.append(sheet.cell(row=i, column=j).value)

            # print(type(row[0]), row)
            if row[0] :
                tmp= replace_fun(row[5])
                row[5] = tmp

                if row[6] and row[6].find('接外宾') > -1:
                    row[6] = '接客户'

                if row[3] != None and row[3].find('商务车') > 0:
                    row[3] = row[3][:-5]
                if row[7].find('小车') > -1:
                    row[7] = '小车'
                else:
                    row[7] = '商务车'
                row_0 = str(row[0])
                row[0] = row_0

                if row_0.find('日期') > -1:
                    print("首行", row)
                elif row_0.find('车费') < 0:
                    # if row[0] == '2017.05.26':
                    #     print(row)

                    if sheet_name.find('商务')>-1:
                        sw.append(row)
                        active_num += 1
                    elif sheet_name.find('小车')>-1:
                        xc.append(row)
                        active_num += 1
                    elif sheet_name.find('自由')>-1:
                        zy.append(row)
                        active_num += 1
                # else:
                #     print("ddddd", row)
                # else:
                #     if sheet_name.find('商务')>-1:
                #         sw.append(row)
                #     elif sheet_name.find('小车')>-1:
                #         xc.append(row)
                #     elif sheet_name.find('自由')>-1:
                #         zy.append(row)
        print("有效行数", active_num)
    wb.close()
    return sw, xc, zy
list_chae = []

shangwu_l = []
xiaoche_l = []
freespace_l = []

sum_= [0,0,0]

new_freespace = []
new_professor = []

for file in file_list_can:
    sw, xc, zy = read_file(file)
    shangwu_l.extend(sw)
    xiaoche_l.extend(xc)
    freespace_l.extend(zy)

# print(shangwu_l)
name_idx = 5

def gen_list(src_list):
    for row in src_list:
        row[0] = replace_time(row[0])
        if row[name_idx] in professor_list:
            new_professor.append(row)
        else:
            new_row = row.copy()
            if row[2] != None and row[7].find('小车') < 0:

                if row[name_idx - 1] > 140 and row[name_idx - 1] < 261:
                    if int(row[2]) < 3:
                        row[7] = '小车'
                        new_row[7] = '小车'
                        new_row[name_idx - 1] -= 140
                        # print('出现核减',new_row)
                        list_chae.append(new_row)
                        row[name_idx - 1] = 140
                        # print(row)

            # if row[2] != None and int(row[2]) > 4:
            #     row[2] = 4
            new_freespace.append(row)


def gen_dic(key_idx, data_l):
    dic = {}
    for row in data_l:
        if key_idx == 0:
            key = row[key_idx][:row[key_idx].rfind('.')]
        else:
            key = row[key_idx]
        if key in dic.keys():
            dic[key].append(row)
        else:
            dic[key] = [row]

    return dic

print("自由空间清单")
gen_list(freespace_l)
print("小车清单")
gen_list(xiaoche_l)
print("商务清单")
gen_list(shangwu_l)

def add_chepai(list_in, type_lst):
    for row in list_in:
        row[9] = chepai[type_lst(random.randint(0,len(type_lst)-1))]

# def add_chepai(list_in):
#     sorted(list_in, key=lambda times: times[0])
#     for row in list_in:
#         row[8] = chepai[random.randint(0, 4)]
#
# def div_type(list_in):
#     list1 = [[],[]]
#
#     for row in list_in:
#         list1[row[9]].append(row)


dic_sw = gen_dic(name_idx, new_professor)
dic_zy = gen_dic(0, new_freespace)



def gen_results(dic, path, name_idx):
    wb = Workbook()
    total = []
    for key in sorted(dic.keys()):
        sum_ = 0
        sum_hejian = 0


        # print(key, len(dic[key]))
        new_l = sorted(dic[key], key=lambda times: times[0])
        # print(dic[key])
        ws = wb.create_sheet(title=key)

        for col in range(1,len(head_title)+1):
            ws.cell(column=col, row=1, value=head_title[col-1])

        for row in range(1, len(new_l)+1):
            if new_l[row-1][7].find('小车') < 0:
                new_l[row - 1][8] = chepai_sw[random.randint(0, 1)]
            else:
                new_l[row - 1][8] = chepai[random.randint(0,4)]
            # dic[key][row][7] = new_l[row - 1][8]
            for col in range(1,len(new_l[row-1])+1):
                ws.cell(column=col, row=row+1, value=new_l[row-1][col-1])
            tmp_col = len(new_l[row-1])+2
            tmp_val = new_l[row-1][name_idx-1]

            # print(dic[key][row-1])
            # if dic[key][row-1][2] != None and dic[key][row-1][7].find('小车') < 0:
            #     if tmp_val > 140 and tmp_val < 261:
            #         if int(dic[key][row-1][2]) < 3:
            #
            #             #####差额放到另外一张表里
            #             tmp_val = 140
            # ws.cell(column=9, row=row+1, value=tmp_val)
            sum_hejian += tmp_val

            sum_ += dic[key][row-1][name_idx-1]
        ws.cell(column=1, row=len(dic[key])+3, value="车费(元）")
        ws.cell(column=2, row=len(dic[key]) + 3, value=sum_)
        ws.cell(column=3, row=len(dic[key]) + 3, value='税（车费*5%）：')
        ws.cell(column=4, row=len(dic[key]) + 3, value=sum_*0.05)
        ws.cell(column=5, row=len(dic[key]) + 3, value='含税总计')
        ws.cell(column=6, row=len(dic[key]) + 3, value=sum_*1.05)

        # ws.cell(column=1, row=len(dic[key]) + 4, value="核减后车费(元）")
        # ws.cell(column=2, row=len(dic[key]) + 4, value=sum_hejian)
        # ws.cell(column=3, row=len(dic[key]) + 4, value='税（车费*5%）：')
        # ws.cell(column=4, row=len(dic[key]) + 4, value=sum_hejian * 0.05)
        # ws.cell(column=5, row=len(dic[key]) + 4, value='核减后含税总计')
        # ws.cell(column=6, row=len(dic[key]) + 4, value=sum_hejian * 1.05)
        total.append([key, sum_, sum_*1.05])
    ws = wb.get_sheet_by_name('Sheet')
    total_sum = 0
    total_sum_hejian = 0
    for row in range(len(total)):
        for col in range(len(total[row])):
            ws.cell(column=col+1, row=row+2, value=total[row][col])
        total_sum += total[row][1]
        # total_sum_hejian += total[row][3]


    ws.cell(column=1, row=1, value='项目')
    ws.cell(column=2, row=1, value='总计')
    ws.cell(column=3, row=1, value='含税总计')
    # ws.cell(column=4, row=1, value='核减后总计')
    # ws.cell(column=5, row=1, value='核减后含税总计')

    ws.cell(column=1, row=len(total) + 3, value='总计')
    ws.cell(column=2, row=len(total) + 3, value=total_sum)
    ws.cell(column=3, row=len(total) + 3, value=total_sum*1.05)
    # ws.cell(column=4, row=len(total) + 3, value=total_sum_hejian)
    # ws.cell(column=5, row=len(total) + 3, value=total_sum_hejian*1.05)

    wb.save(path)
    wb.close()
    return total_sum

def gen_results_person(dic, path, name_idx):

    for key in sorted(dic.keys()):
        wb = Workbook()
        sum_ = 0
        print('教授个人用车情况')
        print(key, len(dic[key]))
        # print(dic[key])
        new_l = sorted(dic[key], key=lambda times: times[0])
        ws = wb.get_sheet_by_name('Sheet')

        for col in range(1,len(head_title)+1):
            ws.cell(column=col, row=1, value=head_title[col-1])

        for row in range(1, len(new_l)+1):
            for col in range(1,len(new_l[row-1])+1):
                ws.cell(column=col, row=row+1, value=new_l[row-1][col-1])

            sum_ += new_l[row-1][name_idx-1]
        ws.cell(column=1, row=len(dic[key])+3, value="车费(元）")
        ws.cell(column=2, row=len(dic[key]) + 3, value=sum_)
        ws.cell(column=3, row=len(dic[key]) + 3, value='税（车费*5%）：')
        ws.cell(column=4, row=len(dic[key]) + 3, value=sum_*0.05)
        ws.cell(column=5, row=len(dic[key]) + 3, value='含税总计')
        ws.cell(column=6, row=len(dic[key]) + 3, value=sum_*1.05)

        new_path = path.split('.')[0]+'-'+key+'.'+path.split('.')[1]
        wb.save(new_path)
        wb.close()

total_money += gen_results(dic_zy, result_zy_file_can, name_idx)
total_money += gen_results(dic_sw, result_sw_file_can, name_idx)
gen_results_person(dic_sw, result_sw_file_can, name_idx)

#单独额外处理陈宇翱
list_cya = dic_sw['陈宇翱']
print('陈宇翱用车：')
print(list_cya)
dic_cya = gen_dic(0, list_cya)
gen_results(dic_cya, result_cya_file_can, name_idx)



############################################can not
shangwu_l = []
xiaoche_l = []
freespace_l = []

sum_= [0,0,0]

new_freespace = []
new_professor = []

for file in file_list_canot:
    sw, xc, zy = read_file(file)
    shangwu_l.extend(sw)
    xiaoche_l.extend(xc)
    freespace_l.extend(zy)

name_idx = 5

print("自由空间清单")
gen_list(freespace_l)
print("小车清单")
gen_list(xiaoche_l)
print("商务清单")
gen_list(shangwu_l)

dic_sw = gen_dic(name_idx, new_professor)
dic_zy = gen_dic(0, new_freespace)

total_money += gen_results(dic_zy, result_zy_file, name_idx)
total_money += gen_results(dic_sw, result_sw_file, name_idx)
gen_results_person(dic_sw, result_sw_file, name_idx)


#单独额外处理陈宇翱
list_cya = dic_sw['陈宇翱']
print('陈宇翱用车：')
print(list_cya)
dic_cya = gen_dic(0, list_cya)
gen_results(dic_cya, result_cya_file, name_idx)

dic_ce = gen_dic(0, list_chae)
total_money += gen_results(dic_ce, result_ce_file, name_idx)



print("总金额：", total_money)

#
# print("自由空间清单")
# for row in freespace_l:
#     if row[name_idx] in professor_list:
#         new_professor.append(row)
#     else:
#         new_freespace.append(row)
#
#
# print("小车清单")
# for row in xiaoche_l:
#     if row[name_idx] in professor_list:
#         new_professor.append(row)
#     else:
#         new_freespace.append(row)
#
# print("商务车清单")
# for row in shangwu_l:
#     if row[name_idx] in professor_list:
#         new_professor.append(row)
#     else:
#         new_freespace.append(row)